import openpyxl 
from openpyxl.utils import range_boundaries
wbook=openpyxl.load_workbook("./test/openpyxl_merge_unmerge.xlsx")
sheet=wbook["unmerge_sample"]
for cell_group in sheet.merged_cells.ranges:
	min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
	top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
	sheet.unmerge_cells(str(cell_group))
	for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
		for cell in row:
			cell.value = top_left_cell_value
wbook.save("./test/openpyxl_merge_unmerge.xlsx")


