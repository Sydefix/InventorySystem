
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

wb = load_workbook('MC-40 22C01.xlsx')
sheet = wb['sheet1']

loadimage = SheetImageLoader(sheet)


for anycategory in sheet.iter_cols(min_row=2, min_col=2, max_col=2, max_row=300):
    for cell in anycategory:
        if cell.value is None:
            continue
        print("processing: "+cell.value + "--> " + cell.coordinate)
        if loadimage.image_in(cell.coordinate.replace("B", "A")):
            #error if there is no folder
            tmp = loadimage.get(cell.coordinate.replace("B", "A")).save("assets/"+ cell.value + ".png")
            print("Got it!")

