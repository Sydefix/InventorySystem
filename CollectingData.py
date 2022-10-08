

from openpyxl import load_workbook
from collections import defaultdict
import json

wb = load_workbook('MC-40 22C01.xlsx')
sheet = wb['sheet1']

royal_dict = defaultdict(lambda : defaultdict(dict))

def getvalue(sheet, col): 
    return sheet[cell.coordinate.replace('B',col)].value

current_cell: str = ""


for anycategory in sheet.iter_cols(min_row=2, min_col=2, max_col=2, max_row=139):
    for cell in anycategory:

        if cell.value is not None:
            current_cell = cell.value
        
        getvalue = lambda col: sheet[cell.coordinate.replace('B',col)].value 
        print(f"processing: {current_cell} <---> {getvalue('C')}")
        
        royal_dict[current_cell][getvalue('C')]["QuantityPerCTN"] = getvalue("D")
        royal_dict[current_cell][getvalue('C')]["CTN"] = getvalue("E")
        royal_dict[current_cell][getvalue('C')]["RMB"] = getvalue("F")
        royal_dict[current_cell][getvalue('C')]["Douane"] = getvalue("G")
        royal_dict[current_cell][getvalue('C')]["Transport"] = getvalue("H")
        royal_dict[current_cell][getvalue('C')]["Total"] = getvalue("I")


final_result = json.dumps(royal_dict, indent=4)
# print(final_result)

# Serializing json
final_result = json.dumps(royal_dict, indent=4)

# Writing to sample.json
with open("MC-40 22C01.json", "w") as outfile:
    outfile.write(final_result)

